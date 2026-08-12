"""
공공도서관 도서 검색 프로그램 (전국 광역단위 / data 디렉터리 통합 지원)

실행 흐름:
  1. cache.txt 에서 이전에 선택한 지역/도서관 목록을 읽음
  2. 사용자에게 이 지역을 그대로 사용할지 대화박스로 질문
     - No → 광역단체 선택 → 하위 지역 선택 → cache.txt 갱신
  3. 선택된 지역별 엑셀 파일을 엑셀로 열고, 사용자가 편집/닫을 때까지 대기
  4. 저장된 엑셀 내용에 따라 도서관 검색 수행
  5. 콘솔에 결과 출력 후 엑셀 파일을 다시 열어 사용자에게 보여줌
  6. 프로그램 종료
"""
import os
import sys
import re
import subprocess
import time
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

from scrapers import get_scraper, get_metro_map, find_metro_by_region, METRO_MAP
from title_utils import strip_title
from gui_utils import check_file_writable, show_error_and_exit

import tkinter as tk
from tkinter import messagebox

# PyInstaller one-file 빌드 시 exe 위치 기준, 일반 실행 시 스크립트 위치 기준
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
CACHE_FILE = BASE_DIR / "cache.txt"


# ─── cache.txt 입출력 ────────────────────────────────────────────

def read_cache() -> list[dict]:
    """
    cache.txt를 읽어 [{"metro": "경기도", "region": "화성시"}, ...] 형태로 반환.
    파일이 없거나 비어 있으면 빈 리스트 반환.
    """
    if not CACHE_FILE.exists():
        return []

    entries = []
    try:
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split("/", maxsplit=1)
                if len(parts) == 2:
                    entries.append({"metro": parts[0].strip(), "region": parts[1].strip()})
    except Exception:
        return []
    return entries


def write_cache(entries: list[dict]):
    """cache.txt에 선택된 지역 목록을 기록합니다."""
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        f.write("# 선택된 광역단체/지역 목록 (자동 생성)\n")
        for e in entries:
            f.write(f"{e['metro']}/{e['region']}\n")


# ─── GUI 다이얼로그 ──────────────────────────────────────────────

def _create_hidden_root() -> tk.Tk:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    return root


def ask_yes_no(title: str, message: str) -> bool:
    """Yes/No 대화박스. True=Yes, False=No."""
    root = _create_hidden_root()
    result = messagebox.askyesno(title, message, parent=root)
    root.destroy()
    return result


def ask_multi_select(title: str, items: list[str]) -> list[str]:
    """
    체크박스 리스트 다이얼로그로 여러 항목 선택.
    OK 시 선택된 항목 리스트 반환, 취소 시 빈 리스트.
    """
    root = tk.Tk()
    root.title(title)
    root.attributes("-topmost", True)
    root.resizable(False, False)

    selected: list[str] = []

    # ── 스크롤 가능 프레임 ──
    outer = tk.Frame(root)
    outer.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)

    canvas = tk.Canvas(outer, width=340, height=min(len(items) * 28 + 20, 460))
    scrollbar = tk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
    inner = tk.Frame(canvas)

    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=inner, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # 마우스 휠 스크롤
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    canvas.bind_all("<MouseWheel>", _on_mousewheel)

    vars_map: dict[str, tk.BooleanVar] = {}
    for item in items:
        var = tk.BooleanVar(value=False)
        vars_map[item] = var
        cb = tk.Checkbutton(inner, text=item, variable=var, anchor="w", font=("맑은 고딕", 10))
        cb.pack(fill=tk.X, padx=6, pady=1)

    # ── 전체선택/해제 + OK/Cancel 버튼 ──
    btn_frame = tk.Frame(root)
    btn_frame.pack(fill=tk.X, padx=8, pady=6)

    def toggle_all():
        any_on = any(v.get() for v in vars_map.values())
        for v in vars_map.values():
            v.set(not any_on)

    def on_ok():
        nonlocal selected
        selected = [k for k, v in vars_map.items() if v.get()]
        root.destroy()

    def on_cancel():
        root.destroy()

    tk.Button(btn_frame, text="전체선택/해제", command=toggle_all, width=14).pack(side=tk.LEFT, padx=2)
    tk.Button(btn_frame, text="확인", command=on_ok, width=10).pack(side=tk.RIGHT, padx=2)
    tk.Button(btn_frame, text="취소", command=on_cancel, width=10).pack(side=tk.RIGHT, padx=2)

    root.update_idletasks()
    w = root.winfo_width()
    h = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (w // 2)
    y = (root.winfo_screenheight() // 2) - (h // 2)
    root.geometry(f"+{x}+{y}")

    root.mainloop()
    return selected


# ─── 지역 선택 흐름 ──────────────────────────────────────────────

def select_regions_interactively() -> list[dict]:
    """대화박스를 통해 광역단체 → 하위 지역 선택 후 결과 반환."""
    metro_names = list(METRO_MAP.keys())

    # 1단계: 광역단체 선택
    chosen_metros = ask_multi_select("① 광역단체 선택", metro_names)
    if not chosen_metros:
        print("  [안내] 광역단체가 선택되지 않았습니다. 프로그램을 종료합니다.")
        sys.exit(0)

    # 2단계: 각 광역단체 내 지역 선택
    entries: list[dict] = []
    for metro in chosen_metros:
        regions = METRO_MAP.get(metro, [])
        if not regions:
            continue
        chosen_regions = ask_multi_select(f"② [{metro}] 지역/도서관 선택", regions)
        for r in chosen_regions:
            entries.append({"metro": metro, "region": r})

    if not entries:
        print("  [안내] 지역이 선택되지 않았습니다. 프로그램을 종료합니다.")
        sys.exit(0)

    return entries


# ─── 엑셀 열기/대기 ──────────────────────────────────────────────

def open_excel_and_wait(file_path: str):
    """
    엑셀 파일을 OS 기본 프로그램으로 열고,
    사용자가 닫을 때까지(=파일 잠금이 풀릴 때까지) 대기합니다.
    """
    path = Path(file_path)
    if not path.exists():
        print(f"  [경고] 파일 없음: {file_path}")
        return

    # 엑셀 열기
    os.startfile(str(path))
    print(f"\n  [열기] 엑셀 파일을 열었습니다: {path.name}")
    print("     편집 후 엑셀을 저장하고 닫아주세요. 닫힐 때까지 대기합니다...")

    # 엑셀이 파일을 잠글 때까지 잠깐 대기
    time.sleep(2)

    # 파일 잠금이 풀릴 때까지 폴링
    while True:
        try:
            with open(str(path), "r+b") as f:
                pass
            # 잠금이 풀렸으면 루프 탈출
            break
        except (PermissionError, OSError):
            time.sleep(1)

    print("  [확인] 엑셀 파일이 닫혔습니다. 검색을 시작합니다.\n")


def open_excel_readonly(file_path: str):
    """결과 확인용으로 엑셀 파일을 열기만 합니다 (대기 안 함)."""
    path = Path(file_path)
    if path.exists():
        os.startfile(str(path))


# ─── 엑셀 처리/검색 ─────────────────────────────────────────────

def sanitize_sheet_name(name: str) -> str:
    """엑셀 시트 이름으로 사용할 수 없는 문자를 제거하고 길이 제한(31자)을 적용합니다."""
    if not name:
        return "기타도서관"
    clean_name = re.sub(r'[\\/*?:\[\]]', ' ', name).strip()
    return clean_name[:31] if clean_name else "기타도서관"


def is_header(val_a: str, val_b: str) -> bool:
    """첫 번째 행이 헤더인지 여부를 판단합니다."""
    headers_a = ["책 정보", "책정보", "책 이름", "책이름", "책 제목", "제목", "title", "book title"]
    headers_b = ["저자", "author"]
    val_a_lower = val_a.strip().lower()
    val_b_lower = val_b.strip().lower()
    return val_a_lower in headers_a or val_b_lower in headers_b


def ensure_excel_template(metro: str, region: str) -> Path:
    """지역 엑셀 파일이 없으면 템플릿을 생성하고, 파일 경로를 반환합니다."""
    metro_dir = DATA_DIR / metro
    metro_dir.mkdir(parents=True, exist_ok=True)

    file_path = metro_dir / f"{region}.xlsx"
    if not file_path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "입력"
        ws.append(["책 이름", "저자"])
        wb.save(file_path)
    return file_path


def process_excel_file(file_path: str):
    """엑셀 파일을 읽어 도서관 검색 후 결과를 시트에 기록합니다."""
    path = Path(file_path)
    file_name = path.name
    region_name = path.stem.strip()

    # 파일 접근 및 락 여부 검사
    check_file_writable(str(path))

    # 상위 디렉터리에서 광역단위 이름 판별
    parent_dir_name = path.parent.name
    metro_map = get_metro_map()

    if parent_dir_name in metro_map:
        metro_name = parent_dir_name
    else:
        metro_name = find_metro_by_region(region_name)

    print(f"\n{'='*65}")
    print(f"  [광역단위] : {metro_name}")
    print(f"  [처리 파일] : {file_name}")
    print(f"  [지역/기관] : {region_name}")
    print(f"{'='*65}")

    # 엑셀 워크북 열기
    try:
        wb = openpyxl.load_workbook(str(path))
    except PermissionError:
        msg = (
            f"'{file_name}' 파일에 접근할 수 없습니다.\n"
            f"현재 열려 있는 엑셀(Excel) 프로그램을 닫고 다시 실행해 주세요."
        )
        show_error_and_exit("엑셀 파일 접근 오류", msg)
        return
    except Exception as e:
        print(f"  [오류] 엑셀 파일 열기 실패: {e}")
        return

    if '입력' not in wb.sheetnames:
        print(f"  [경고] '{file_name}' 파일에 '입력' 시트가 없습니다. 스킵합니다.")
        return

    sheet_in = wb['입력']

    # 입력 시트 데이터 읽기
    input_items = []
    for row_idx, row in enumerate(sheet_in.iter_rows(values_only=True), 1):
        if not row:
            continue
        val_a = str(row[0]).strip() if row[0] is not None else ""
        val_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        if not val_a:
            continue

        # 헤더 행 스킵
        if row_idx == 1 and is_header(val_a, val_b):
            continue

        input_items.append((val_a, val_b))

    if not input_items:
        print("  [알림] '입력' 시트에 검색할 책 데이터가 없습니다.")
        return

    print(f"  >> 총 {len(input_items)}건의 검색 대상을 읽었습니다.")

    # 검색 시작 전에 기존 출력 시트 삭제 ('입력' 시트만 남김)
    for sname in list(wb.sheetnames):
        if sname != '입력':
            del wb[sname]

    # 스크래퍼 구하기 및 데이터 수집
    scraper = get_scraper(region_name, metro_name)
    library_books = defaultdict(list)

    for raw_title, author in input_items:
        search_title = strip_title(raw_title)
        print(f"\n  [검색] 원제목: '{raw_title}' -> 검색어: '{search_title}'" + (f", 저자: '{author}'" if author else ""))

        try:
            total_cnt, books = scraper.search(search_title, author)
            print(f"  >> 검색 결과: 총 {total_cnt}건")

            for book in books:
                lib_name = book.library.strip() if book.library else f"{region_name}도서관"

                # 스마트도서관, 작은도서관 제외
                if "스마트" in lib_name or "작은" in lib_name:
                    continue

                library_books[lib_name].append(book)
        except Exception as e:
            print(f"  [오류] 검색 중 문제 발생: {e}")

    if not library_books:
        print("\n  [알림] 검색된 (조건에 맞는) 도서 결과가 없습니다.")

    # 각 도서관 이름으로 시트 생성 및 데이터 출력
    headers = ["책 이름", "청구기호", "음원 여부", "저자", "도서관 내 열람실"]

    for lib_name, books in library_books.items():
        sheet_name = sanitize_sheet_name(lib_name)

        if sheet_name in wb.sheetnames:
            sheet_out = wb[sheet_name]
            sheet_out.delete_rows(1, sheet_out.max_row + 1)
        else:
            sheet_out = wb.create_sheet(title=sheet_name)

        sheet_out.append(headers)

        for book in books:
            sheet_out.append([
                book.title or "",
                book.call_number or "",
                book.has_audio or "",
                book.author or "",
                book.location or ""
            ])

        # 열 너비 자동 조정
        for col in sheet_out.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                cell_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                if cell_len > max_len:
                    max_len = cell_len
            col_letter = get_column_letter(col[0].column)
            sheet_out.column_dimensions[col_letter].width = max(max_len + 3, 12)

        print(f"  [시트 생성] '{sheet_name}' (도서 {len(books)}건)")

    # 엑셀 파일 저장
    check_file_writable(str(path))

    try:
        wb.save(str(path))
        print(f"\n  >> '{path}' 저장 완료!")
    except PermissionError:
        msg = (
            f"'{file_name}' 저장 단계에서 접근 권한이 거부되었습니다.\n"
            f"열려있는 엑셀(Excel) 프로그램을 완전히 닫고 다시 실행해 주세요."
        )
        show_error_and_exit("엑셀 저장 실패 (Permission Denied)", msg)
    except Exception as e:
        print(f"\n  [오류] 엑셀 파일 저장 실패: {e}")

    # 콘솔에 검색 결과 요약 출력
    print(f"\n{'─'*65}")
    print(f"  [결과] [{region_name}] 검색 결과 요약")
    print(f"{'─'*65}")
    if library_books:
        for lib_name, books in library_books.items():
            print(f"    - {lib_name}: {len(books)}건")
        total_books = sum(len(b) for b in library_books.values())
        print(f"\n    총 {len(library_books)}개 도서관, {total_books}건의 도서")
    else:
        print("    (검색 결과 없음)")
    print(f"{'─'*65}")


# ─── 메인 흐름 ───────────────────────────────────────────────────

def main():
    print("\n" + "=" * 65)
    print("  [전국 공공도서관 도서 검색 프로그램]")
    print("=" * 65)

    # ── 1.1 cache.txt 읽기 ──
    cached = read_cache()

    # ── 1.2 cache.txt 가 있으면 사용자에게 확인 ──
    entries: list[dict] = []

    if cached:
        region_display = ", ".join(f"{e['region']}({e['metro']})" for e in cached)
        use_cache = ask_yes_no(
            "이전 선택 지역 확인",
            f"이전에 선택한 지역이 있습니다:\n\n{region_display}\n\n이 지역으로 계속하시겠습니까?"
        )
        if use_cache:
            entries = cached
        else:
            # ── 1.2.1 '아니오' 선택 시 기존 cache.txt 삭제 ──
            if CACHE_FILE.exists():
                try:
                    CACHE_FILE.unlink()
                except Exception:
                    pass
            entries = select_regions_interactively()
            write_cache(entries)
    else:
        # cache.txt 없음 → 새로 선택
        entries = select_regions_interactively()
        write_cache(entries)

    if not entries:
        print("  [안내] 선택된 지역이 없습니다. 프로그램을 종료합니다.")
        return

    # 선택된 지역 출력
    print("\n  ▶ 선택된 지역:")
    for e in entries:
        print(f"    • {e['metro']} / {e['region']}")

    # ── 1.3 선택된 지역의 엑셀들을 열고 편집 대기 ──
    excel_files: list[str] = []
    for e in entries:
        fp = ensure_excel_template(e["metro"], e["region"])
        excel_files.append(str(fp))

    # 모든 엑셀 파일을 열기
    for fp in excel_files:
        os.startfile(fp)

    print(f"\n  [열기] {len(excel_files)}개 엑셀 파일을 열었습니다.")
    print("     각 엑셀의 '입력' 시트에 검색할 책 정보를 입력하세요.")
    print("     모든 엑셀을 저장하고 닫으면 검색이 시작됩니다.\n")

    # 엑셀이 파일을 잠글 때까지 잠깐 대기
    time.sleep(3)

    # 모든 엑셀 파일의 잠금이 풀릴 때까지 대기
    for fp in excel_files:
        while True:
            try:
                with open(fp, "r+b") as f:
                    pass
                break
            except (PermissionError, OSError):
                time.sleep(1)

    print("  [확인] 모든 엑셀 파일이 닫혔습니다. 검색을 시작합니다.\n")

    # ── 1.4 저장된 엑셀 내용에 따라 도서관 검색 ──
    for fp in excel_files:
        process_excel_file(fp)

    # ── 1.5 검색 완료 후 엑셀 다시 열기 ──
    print(f"\n{'='*65}")
    print("  [완료] 모든 검색이 완료되었습니다! 결과 엑셀을 엽니다.")
    print(f"{'='*65}\n")

    for fp in excel_files:
        open_excel_readonly(fp)

    # ── 1.6 프로그램 종료 ──
    print("  프로그램을 종료합니다.")


if __name__ == "__main__":
    main()
