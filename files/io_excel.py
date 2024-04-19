import os
import openpyxl
from openpyxl.styles import Color
from openpyxl.styles import PatternFill

from .io_file import *

idx_fmt_cell = '%s%d'
# need to care everytime the order of values per record changes
list_hl_width_idx = ['A', 'D']
list_hl_width_val = 30
list_hl_kwd_color_idx = ['A', 'C']
list_hl_val_color_idx = ['D', 'E']
kwd_color = PatternFill(fill_type='solid', fgColor=Color('eaeef3'))
val_color = PatternFill(fill_type='solid', fgColor=Color('eae0B6'))


def getCellIdx(col, row):
    return idx_fmt_cell % (col, row)


class IoExcel(IoFile):
    @staticmethod
    def create(path_in, path_out, list_libs):
        obj = None
        ext_in = os.path.splitext(path_in)[1]
        #print("%s" % ext_in)
        if ext_in == '.xlsx' or ext_in == '.xls':
            obj = IoExcel(path_in, path_out, list_libs)
        return obj

    def __init__(self, path_in, path_out, list_libs):
        self.path_in = path_in
        self.path_out = path_in     # for tabbed files
        self.wb = openpyxl.load_workbook(path_in)
        self.list_libs = list_libs
        self.ws = None
        self.num_of_kwds = 0
        self.col_base = ord('A')
        self.col_ofs = 0
        self.row = 0
        self.iter_dir = IterableGetDirName(self.wb)
        super(__class__, self).__init__()

    def get_dir_iter(self):
        try:
            res = self.iter_dir.__next__()
        except StopIteration:
            res = ''
        self.ws = self.wb[res]
        self.col_ofs = 0
        return res

    def set_dir(self, name):
        self.ws = self.wb[name]

    def create_values(self, key):
        # print("%s: %s %d" % (key, chr(self.col_base), self.col_ofs))
        self.iter_get[key] = IterableGetValue(self.ws, chr(self.col_base + self.col_ofs))
        self.col_ofs += 1

    def init_result(self, key):
        ws = None
        try:
            ws = self.wb[key]
            self.wb.remove(ws)
        except KeyError:
            print("%s sheet doesn't exist, so let's create it." % key)
        finally:
            self.ws = self.wb.create_sheet(key)
        # need to care everytime the order of values per record changes
        for idx in list_hl_width_idx:
            # print("%s: %d" % (idx, list_hl_width_val)) ###
            self.ws.column_dimensions[idx].width = list_hl_width_val

    def set_color(self, _list, color, row):
        for idx in _list:
            cell = self.ws[getCellIdx(idx, row)]
            cell.fill = color

    def create_keyword_iter(self, row):
        self.row = row
        self.num_of_kwds = 0
        self.iter_set = IterableSetValue(self.ws, self.col_base)
        self.set_color(list_hl_kwd_color_idx, kwd_color, row)

    def insert_keyword_iter(self, val):
        self.iter_set.set_value(val, self.row)
        self.num_of_kwds += 1

    def create_value_iter(self, row):
        self.row = row
        col = self.col_base + self.num_of_kwds
        self.iter_set = IterableSetValue(self.ws, col)
        self.set_color(list_hl_val_color_idx, val_color, row)

    def insert_value_iter(self, val):
        self.iter_set.set_value(val, self.row)

    def __del__(self):
        for idx in list_hl_width_idx:
            self.ws.column_dimensions[idx].width = list_hl_width_val
        self.wb.save(self.path_in)
        self.wb.close()

class IterableGetDirName(Iterable):
    def __init__(self, wb):
        self.names = wb.sheetnames
        self.len = len(wb.sheetnames)
        self.idx = 0
        # super(__class__, self).__init__()

    def __next__(self):
        if self.idx >= self.len:
            raise StopIteration
        else:
            res = self.names[self.idx]
            self.idx += 1
        return res


class IterableGetValue(Iterable):
    def __init__(self, ws, col):
        self.ws = ws
        self.col = col
        self.idx = 1

    def __next__(self):
        # print("%s %d" % (self.col, self.idx))
        res = self.ws[getCellIdx(self.col, self.idx)].value
        if not isinstance(res, str):
            res = ''
        self.idx += 1
        return res


class IterableSetValue(Iterable):
    def __init__(self, ws, col):
        self.ws = ws
        self.col = col
        self.idx = 0

    def __next__(self):
        pass

    def set_value(self, val, row):
        self.ws[getCellIdx(chr(self.col), row)].value = val
        self.col = self.col + 1
