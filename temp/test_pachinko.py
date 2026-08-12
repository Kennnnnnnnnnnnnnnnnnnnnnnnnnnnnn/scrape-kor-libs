"""
화성시, 부산도서관 대상 '파친코' 검색 비대화형 테스트
(GUI 대화박스 없이 검색 로직만 검증)
"""
import os
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

from scrapers import get_scraper, find_metro_by_region
from title_utils import strip_title

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"


def prepare_test_excel(metro: str, region: str, books: list[tuple[str, str]]) -> Path:
    """테스트용 엑셀 생성 ('입력' 시트에 검색할 책 기록)"""
    metro_dir = DATA_DIR / metro
    metro_dir.mkdir(parents=True, exist_ok=True)
    file_path = metro_dir / f"{region}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "입력"
    ws.append(["책 이름", "저자"])
    for title, author in books:
        ws.append([title, author])
    wb.save(file_path)
    return file_path


def process_and_print(file_path: Path, metro: str, region: str):
    """엑셀 파일 읽고 검색 후 결과 저장 및 콘솔 출력"""
    import re

    wb = openpyxl.load_workbook(str(file_path))
    if '입력' not in wb.sheetnames:
        print(f"  [경고] '입력' 시트 없음: {file_path.name}")
        return

    sheet_in = wb['입력']
    input_items = []
    for row_idx, row in enumerate(sheet_in.iter_rows(values_only=True), 1):
        if not row:
            continue
        val_a = str(row[0]).strip() if row[0] is not None else ""
        val_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not val_a:
            continue
        headers_a = ["책 정보", "책정보", "책 이름", "책이름", "책 제목", "제목", "title", "book title"]
        headers_b = ["저자", "author"]
        if row_idx == 1 and (val_a.strip().lower() in headers_a or val_b.strip().lower() in headers_b):
            continue
        input_items.append((val_a, val_b))

    if not input_items:
        print("  [알림] 검색할 책 데이터 없음")
        return

    print(f"\n{'='*65}")
    print(f"  [광역단위] : {metro}")
    print(f"  [지역/기관] : {region}")
    print(f"  >> 총 {len(input_items)}건의 검색 대상")
    print(f"{'='*65}")

    scraper = get_scraper(region, metro)
    library_books = defaultdict(list)

    for raw_title, author in input_items:
        search_title = strip_title(raw_title)
        print(f"\n  [검색] 원제목: '{raw_title}' -> 검색어: '{search_title}'" +
              (f", 저자: '{author}'" if author else ""))
        try:
            total_cnt, books = scraper.search(search_title, author)
            print(f"  >> 검색 결과: 총 {total_cnt}건")
            for book in books:
                lib_name = book.library.strip() if book.library else f"{region}도서관"
                if "스마트" in lib_name or "작은" in lib_name:
                    continue
                library_books[lib_name].append(book)
        except Exception as e:
            print(f"  [오류] 검색 중 문제 발생: {e}")

    # 기존 출력 시트 삭제
    for sname in list(wb.sheetnames):
        if sname != '입력':
            del wb[sname]

    # 시트 생성
    headers = ["책 이름", "청구기호", "음원 여부", "저자", "도서관 내 열람실"]
    for lib_name, books in library_books.items():
        clean = re.sub(r'[\\/*?:\[\]]', ' ', lib_name).strip()[:31] or "기타도서관"
        if clean in wb.sheetnames:
            sheet_out = wb[clean]
            sheet_out.delete_rows(1, sheet_out.max_row + 1)
        else:
            sheet_out = wb.create_sheet(title=clean)
        sheet_out.append(headers)
        for book in books:
            sheet_out.append([
                book.title or "", book.call_number or "",
                book.has_audio or "", book.author or "", book.location or ""
            ])
        for col in sheet_out.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                cell_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                if cell_len > max_len:
                    max_len = cell_len
            col_letter = get_column_letter(col[0].column)
            sheet_out.column_dimensions[col_letter].width = max(max_len + 3, 12)
        print(f"  [시트 생성] '{clean}' (도서 {len(books)}건)")

    wb.save(str(file_path))
    print(f"\n  >> '{file_path}' 저장 완료!")

    # 결과 요약
    print(f"\n{'─'*65}")
    print(f"  [결과] [{region}] 검색 결과 요약")
    print(f"{'─'*65}")
    if library_books:
        for lib_name, books in library_books.items():
            print(f"    - {lib_name}: {len(books)}건")
        total_books = sum(len(b) for b in library_books.values())
        print(f"\n    총 {len(library_books)}개 도서관, {total_books}건의 도서")
    else:
        print("    (검색 결과 없음)")
    print(f"{'─'*65}")


def main():
    print("\n" + "=" * 65)
    print("  [TEST] 비대화형 테스트: 화성시, 부산도서관 '파친코' 검색")
    print("=" * 65)

    test_targets = [
        {"metro": "경기도", "region": "화성시"},
        {"metro": "부산광역시", "region": "부산도서관"},
    ]

    search_books = [("파친코", "")]

    for target in test_targets:
        fp = prepare_test_excel(target["metro"], target["region"], search_books)
        process_and_print(fp, target["metro"], target["region"])

    print(f"\n{'='*65}")
    print("  [완료] 모든 테스트 검색이 완료되었습니다!")
    print(f"{'='*65}")


if __name__ == "__main__":
    main()
